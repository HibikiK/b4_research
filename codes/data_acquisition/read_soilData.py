# インポート
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
import datetime
from datetime import date
import linecache
import os
import numpy as np
import h5py
from openpyxl import load_workbook, Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
import math

# datファイルオープン
filepath = './gprMax.dat'
keyword1 = '#material:'
keyword2 = '#waveform:'

with open(filepath) as f:
    lines = f.readlines()
    lines_strip = [line.strip() for line in lines]

# 先頭行と最終行の取得
l_i1 = [i for i, line in enumerate(
    lines_strip) if keyword1 in line]
l_int1 = l_i1[0]

l_i2 = [i for i, line in enumerate(
    lines_strip) if keyword2 in line]
l_int2 = l_i2[0]

start = l_int1 + 1
end = l_int2

# excelファイル作成
if os.path.exists('gprMaterial.xlsx'):
    print('Writing gprMaterial Data')
    os.remove('gprMaterial.xlsx')
else:
    print('Rewriting gprMaterial Data')

wb_new = Workbook()
ws_Mat = wb_new.create_sheet(title='MaterialData')

ws_Mat['B2'] = 'Permittivity'
ws_Mat['C2'] = 'Conductivity'


# 行の内容出力
# permittivity
for a in range(start, end):
    mtData = linecache.getline(filepath, a).strip().split()
    mt_per = mtData[1]
    row_no = a - 3
    ws_Mat[f'B{row_no}'] = mt_per

# conductivity
for b in range(start, end):
    mtData = linecache.getline(filepath, b).strip().split()
    mt_con = mtData[2]
    row_no = b - 3
    ws_Mat[f'C{row_no}'] = mt_con


# ファイル保存
wb_new.move_sheet(ws_Mat, offset=-wb_new.index(ws_Mat))
wb_new.save('gprMaterial.xlsx')
