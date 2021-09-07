# インポート
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
import linecache
import os
import numpy as np
import h5py
from openpyxl import load_workbook, Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt

# datファイルオープン
filepath = './20191209COG_data.txt'

with open(filepath) as f:
    lines = f.readlines()
    lines_strip = [line.strip() for line in lines]

# 全行数の取得
TotalLines = sum([1 for _ in open(filepath)])

# y軸データの取得
print('Trace Number?(1~47):')
inNo = input()
traceNo = int(inNo) - 1
amplitude = []
for a in range(1, TotalLines+1):
    timePoint = linecache.getline(filepath, a).strip().split()
    amp = float(timePoint[traceNo])
    amplitude.append(amp)

# x軸データの取得
filepath1 = 'Time.txt'
f1 = open(filepath1, 'r')
with open(filepath) as r:
    lines = f1.readlines()
    lines_strip = [line.strip() for line in lines]
time = []
for b in range(500):
    times = float(lines_strip[b])
    time.append(times)


# グラフ作成
fig = plt.figure(facecolor='mistyrose')
graphTitle = filepath + ' - TraceNo: ' + inNo
ax = fig.add_subplot(111, title=graphTitle,
                     xlabel='Time[ns]', ylabel='Ez,Field Strength')
ax.grid(c='gainsboro', zorder=1)
ax.scatter(time, amplitude, color='firebrick', marker='o', s=4, zorder=5)

fig.savefig('ScanData.png', facecolor=fig.get_facecolor())


# Excel転記
# excelファイル作成
if os.path.exists('ScanData.xlsx'):
    print('Writing and Drawing ScanData')
    os.remove('ScanData.xlsx')
else:
    print('Writing and Drawing ScanData Data')

wb_new = Workbook()
ws_Scan = wb_new.create_sheet(title='ScanData')

ws_Scan['C2'] = 'Time'
ws_Scan['D2'] = 'Amplitude'
ws_Scan['A1'] = 'Trace No:'
ws_Scan['B1'] = inNo

# 行の内容出力
# permittivity
for c in range(500):
    row_no = c + 3
    ws_Scan[f'C{row_no}'] = time[c]

# conductivity
for d in range(500):
    row_no = d + 3
    ws_Scan[f'D{row_no}'] = amplitude[d]

# グラフ作成1: Pressure Head Fitting
chart = ScatterChart()
# グラフのX軸の範囲を設定する為に、Referenceオブジェクト作る
x_values = Reference(
    ws_Scan, min_col=3, min_row=3, max_row=502)
# Yデータの書き込み
# データの範囲(Y軸)をReferenceで選択
Y = Reference(ws_Scan, min_col=4, min_row=3, max_row=502)
# Seriesオブジェクトを作成
series = Series(Y, x_values, title="Amplitude")
chart.series.append(series)
# デザイン
chart.x_axis.title = 'Time'
chart.y_axis.title = 'Amplitude'
chart.title = 'TraceNo: ' + inNo
chart.height = 8.2
chart.width = 12.6
chart.legend.legendPos = "b"
# A3セルにグラフを表示
ws_Scan.add_chart(chart, "F5")


# ファイル保存
wb_new.move_sheet(ws_Scan, offset=-wb_new.index(ws_Scan))
wb_new.save('ScanData.xlsx')
