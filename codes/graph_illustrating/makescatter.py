# 今回使うモジュールの読み込み
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import numpy as np
import openpyxl

# openpyxlのバージョンを表示
print(openpyxl.__version__)
# 作成するExcelファイル名
filename = "openpyxl_chart.xlsx"
# ワークブックの作成
wb = Workbook()
# ワークシートの作成。第1引数にシート名、第2引数に挿入位置
ws = wb.create_sheet("sheet01", 0)

# データ作成
X = np.arange(0, 100, 5)
Y1 = X
Y2 = 2 * X
Y3 = 3 * X

len_data = len(X)

# セルへの書き込み.row=1でA行、column=1で1列目を選択
# A1セルへ書き込み
ws.cell(row=1, column=1, value="X")
# B1セルへ書き込み
ws.cell(row=1, column=2, value="Y1")
# C1セルへ書き込み
ws.cell(row=1, column=3, value="Y2")
# D1セルへ書き込み
ws.cell(row=1, column=4, value="Y3")

for i in range(0, len_data):
    ws.cell(row=i + 2, column=1, value=X[i])
    ws.cell(row=i + 2, column=2, value=Y1[i])
    ws.cell(row=i + 2, column=3, value=Y2[i])
    ws.cell(row=i + 2, column=4, value=Y3[i])

# ScatterChartオブジェクトを作成
chart = ScatterChart()

num_series = 3
min_row = 2
max_row = min_row + len_data - 1

# グラフのX軸の範囲を設定する為に、Referenceオブジェクト作る
x_values = Reference(ws, min_col=1, min_row=min_row, max_row=max_row)

# データの書き込み
for i in range(0, num_series):
    min_col = i + 2
  # データの範囲(Y軸)をReferenceで選択
    values = Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row)
  # Seriesオブジェクトを作成
    series = Series(values, x_values, title="hoge")
    chart.series.append(series)

# グラフの大きさ
chart.height = 10
chart.width = 10
# グラフの大きさ
chart.title = "Hoge"
# 凡例
chart.legend = None

# 目盛間隔
chart.x_axis.majorUnit = 10
chart.y_axis.majorUnit = 20
# 補助目盛間隔
chart.x_axis.minorUnit = 5
chart.y_axis.minorUnit = 10
# グリッドラインを消す
chart.y_axis.majorGridlines = None
chart.x_axis.majorGridlines = None
#軸の向き(out, corss, inから選ぶ)
chart.x_axis.majorTickMark = "in"
chart.x_axis.minorTickMark = "in"
chart.y_axis.majorTickMark = "in"
chart.y_axis.minorTickMark = "in"

ws.add_chart(chart, "F2")

wb.save('openpyxl_chart.xlsx')
