import numpy as np
from openpyxl import load_workbook, Workbook
import os

print('Copying INSfile Data from "OrganizingInsfile.xlsx"')

# Excelファイルオープン
efilename = 'OrganizingINSfile.xlsx'

wb = load_workbook(efilename)
ws = wb['Sheet1']

# テキストファイル作成
tfilename = 'gprMax_out.INS'
if os.path.exists(tfilename):
    os.remove(tfilename)
else:
    print('Creating textfile')

tf = open('gprMax_out.INS', mode='w')
tf.write('pif $\n')

# データ読み込み
last_row = ws.max_row

L = []
for i in range(2, last_row + 1):
    row_no = i
    line = ws.cell(row_no, 2).value
    index = ws.cell(row_no, 3).value
    column = ws.cell(row_no, 4).value
    L.append(line)
    L.append(' ')
    L.append(index)
    L.append(column)
    L.append('\n')
    s = ''.join(L)

# 文字列貼り付け
tf.write(s)

# テキストファイル保存
tf.close
print('Instruction File Created as "gprMax_out.INS"')
