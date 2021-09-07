import numpy as np
from openpyxl import load_workbook, Workbook
import os

print('Arranging Observed Data in PEST Control File')

# observed data の消去
# テキストファイルオープン
tfilename = 'gprMax.pst'
keyword1 = '* observation data'
keyword2 = '* model command line'

with open(tfilename) as tf:
    lines = tf.readlines()
    lines_strip = [line.strip() for line in lines]

# 削除する行の取得
l_i1 = [i for i, line in enumerate(
    lines_strip) if keyword1 in line]
l_start = l_i1[0] + 2
print(l_start)

l_i2 = [i for i, line in enumerate(
    lines_strip) if keyword2 in line]
l_end = l_i2[0]
print(l_end)

# 削除の実行


# 新しいデータの書き込み
# Excelファイルオープン
efilename = 'OrganizingINSfile.xlsx'

wb = load_workbook(efilename)
ws = wb['Sheet2']

last_row = ws.max_row
L = []
for i in range(1, last_row + 1):
    row_no = i
    index = ws.cell(row_no, 1).value
    value = ws.cell(row_no, 2).value
    weight = ws.cell(row_no, 3).value
    group = ws.cell(row_no, 4).value
    L.append(index)
    L.append('  ')
    L.append(value)
    L.append('  ')
    L.append(weight)
    L.append('  ')
    L.append(group)
    L.append('\n')
    s = ''.join(L)

# 文字列貼り付け
t2filename = 'sample.txt'
if os.path.exists(t2filename):
    os.remove(t2filename)
else:
    print('Creating textfile')
tf2 = open(t2filename, mode='w')
tf2.write(s)

# テキストファイル保存
tf2.close
print('Observed Data in PEST Control File has been Arranged')
